from tempfile import NamedTemporaryFile

import openpyxl
import xlrd
from openpyxl.styles import Font

import settings


class Manager:
    def read(self, file) -> dict:
        """
        Reads data from csv, xls and xlsx files
        Read data is dict with keys in first row,
        and list values as in workbook file rows
        Must be saved or updated in database
        """
        workbook = openpyxl.Workbook()
        if file.mimetype != settings.ALLOWED_MIMETYPES['xlsx']:
            self._convert(file, workbook)
        return self.parse(workbook)

    @staticmethod
    def write(data: list, fn = None, args = (), columns = settings.COLNAMES):
        """
        :param data list of data to write
        :param fn function to call for database
        :param args arguments for function
        :param columns columns matching data headers
        """
        if len(data) == 0:
            return None
        wb = openpyxl.Workbook()
        ws = wb.active
        optimal_widths = []
        tables = [chr(ord('A') + i) for i in range(len(columns))]
        for i, colname in enumerate(columns):
            cell = ws.cell(row = 1, column = i + 1, value = colname)
            cell.font = Font(bold = True)
            for j, item in enumerate(data):
                if fn is not None and colname == "targets":
                    target_names = ''
                    targets_dict_list = fn(*args)
                    for target in targets_dict_list:
                        target_names = ','.join([target['username']])
                    optimal_widths.append(
                        len(target_names) + 1 if isinstance(target_names, str) and len(target_names) > 0 else len(colname))
                    ws.cell(row = j + 2, column = i + 1, value = target_names)
                    continue
                    pass
                value = item.get(colname, '')
                optimal_widths.append(len(value) + 1 if isinstance(value, str) and (0 < len(value) > len(colname)) else len(colname))
                ws.cell(row = j + 2, column = i + 1, value = value)
                pass
            pass
        for (table, width) in zip(tables, optimal_widths):
            ws.column_dimensions[table].width = width
        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        return stream

    @staticmethod
    def parse(workbook) -> dict:
        keys = dict()

        def read_args(*args):
            key = None
            for c in args:
                if c in settings.COLNAMES:
                    key = c
                    keys[key] = []
                    continue
                    pass
                keys[key].append(c)

        sheet = workbook.active
        for col in sheet.iter_cols(values_only = True):
            read_args(*col)
        return keys

    @staticmethod
    def _convert(file, workbook):
        content = file.read()
        xls_book = xlrd.open_workbook(file_contents = content)
        for i in range(0, xls_book.nsheets):
            xls_sheet = xls_book.sheet_by_index(i)
            sheet = workbook.active if i == 0 else workbook.create_sheet()
            sheet.title = xls_sheet.name
            for row in range(0, xls_sheet.nrows):
                for col in range(0, xls_sheet.ncols):
                    sheet.cell(row = row + 1, column = col + 1).value = xls_sheet.cell_value(row, col)
        return workbook

    pass
